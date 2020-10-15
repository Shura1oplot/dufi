# [SublimeLinter @python:3]

import openpyxl
from openpyxl.utils.datetime import CALENDAR_MAC_1904


def read(fname, sheet=None):
    fp = open(fname, "rb")
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)

    datemode = int(wb.excel_base_date == CALENDAR_MAC_1904)

    def _f(ws):
        return ws.title, datemode, _rows_iter(ws)

    if sheet is None:
        yield _f(wb.get_active_sheet())

    elif hasattr(sheet, "match"):
        for name in wb.sheetnames:
            if sheet.match(name):
                yield _f(wb[name])

    elif callable(sheet):
        result = sheet(wb.sheetnames)

        if isinstance(result, (tuple, int)):
            for name in result:
                yield _f(wb[name])

        else:
            yield _f(wb[result])

    elif isinstance(sheet, int):
        yield _f(wb.worksheets(sheet))  # TODO: test

    else:
        yield _f(wb[sheet])


def _rows_iter(ws):
    for row in ws.rows:
        yield [x.value for x in row]


if __name__ == "__main__":
    pass
